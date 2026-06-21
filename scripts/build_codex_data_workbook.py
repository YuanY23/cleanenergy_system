from __future__ import annotations

import json
import math
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "codex数据汇总表.xlsx"

LATITUDE = 39.61
LONGITUDE = 109.78
NASA_URL = (
    "https://power.larc.nasa.gov/api/temporal/hourly/point"
    f"?parameters=ALLSKY_SFC_SW_DWN,WS50M,T2M&community=RE"
    f"&longitude={LONGITUDE}&latitude={LATITUDE}"
    "&start=20240101&end=20241231&format=JSON&time-standard=LST"
)

USD_CNY = 7.25
BASE_ELECTRICITY_PRICE = 0.45
GAS_PRICE = 2.952
GRID_EF_INNER_MONGOLIA = 0.6849
CARBON_PRICE = 62.36


def get_nasa_power() -> dict:
    cache = ROOT / "outputs" / "nasa_power_ordos_2024.json"
    cache.parent.mkdir(exist_ok=True)
    if cache.exists():
        return json.loads(cache.read_text(encoding="utf-8"))
    with urllib.request.urlopen(NASA_URL, timeout=120) as response:
        payload = response.read().decode("utf-8")
    cache.write_text(payload, encoding="utf-8")
    return json.loads(payload)


def pv_capacity_factor(ghi_wh_m2: float, temp_c: float) -> float:
    if ghi_wh_m2 <= 0:
        return 0.0
    module_temp = temp_c + 20.0 * (ghi_wh_m2 / 800.0)
    temp_factor = 1.0 - 0.004 * max(0.0, module_temp - 25.0)
    return round(max(0.0, min(0.95, (ghi_wh_m2 / 1000.0) * 0.82 * temp_factor)), 4)


def wind_capacity_factor(ws50m: float) -> float:
    cut_in, rated, cut_out = 3.0, 12.0, 25.0
    if ws50m < cut_in or ws50m >= cut_out:
        return 0.0
    if ws50m >= rated:
        return 0.95
    return round(0.95 * ((ws50m**3 - cut_in**3) / (rated**3 - cut_in**3)), 4)


def tou_period_and_price(dt: datetime) -> tuple[str, float]:
    h = dt.hour
    low_wind = dt.month in (6, 7, 8)
    if low_wind:
        if 19 <= h < 21:
            return "尖峰", round(BASE_ELECTRICITY_PRICE * 1.54 * 1.20, 4)
        if 13 <= h < 15:
            return "深谷", round(BASE_ELECTRICITY_PRICE * 0.44 * 0.80, 4)
        if 6 <= h < 8 or 18 <= h < 22:
            return "峰", round(BASE_ELECTRICITY_PRICE * 1.54, 4)
        if 11 <= h < 16:
            return "谷", round(BASE_ELECTRICITY_PRICE * 0.44, 4)
        return "平", BASE_ELECTRICITY_PRICE
    if 6 <= h < 8 or 18 <= h < 22:
        return "峰", round(BASE_ELECTRICITY_PRICE * 1.68, 4)
    if 0 <= h < 4 or 11 <= h < 16:
        return "谷", round(BASE_ELECTRICITY_PRICE * 0.48, 4)
    return "平", BASE_ELECTRICITY_PRICE


def season_name(month: int) -> str:
    if month in (12, 1, 2):
        return "冬季"
    if month in (6, 7, 8):
        return "夏季"
    return "过渡季"


def reconstructed_loads(dt: datetime, temp_c: float) -> tuple[float, float, float]:
    h = dt.hour
    month = dt.month
    # Heavy-industry electric load: high baseload, weekday/workshift shape, annual scale
    # calibrated to the public 150 MW peak / 1.2 TWh-y industrial park case.
    shift = 1.0
    if 8 <= h < 18:
        shift = 1.07
    elif 18 <= h < 23:
        shift = 1.03
    elif 0 <= h < 5:
        shift = 0.93
    seasonal = 1.03 if month in (6, 7, 8, 9) else 0.99 if month in (12, 1, 2) else 1.0
    electric = min(150000.0, 136500.0 * shift * seasonal)

    # Heat demand is temperature-sensitive with an industrial process-heat baseload.
    process_heat = 62000.0
    space_heat = max(0.0, 18.0 - temp_c) * 1900.0
    summer_hot_water = 6000.0 if month in (6, 7, 8) else 0.0
    heat = min(112000.0, process_heat + space_heat + summer_hot_water)

    # Hydrogen offtake is a reconstructed industrial/fuel load with daytime logistics uplift.
    h2_base = 960.0
    logistics = 110.0 if 7 <= h < 20 else -45.0
    weekday_like = 35.0 * math.sin((dt.timetuple().tm_yday / 365.0) * 2 * math.pi)
    hydrogen = max(780.0, h2_base + logistics + weekday_like)
    return round(electric, 1), round(heat, 1), round(hydrogen, 2)


def source_rows() -> list[list[object]]:
    return [
        ["S01", "园区背景", "Surbana Jurong: Ordos' new net-zero industrial park", "官方/项目案例", "Ordos, Inner Mongolia; 73 km2; 100% zero-carbon energy supply; hydrogen and air-source heat-pump measures", "https://www.sjgroup.com/projects/ordos/", "2026-06-18"],
        ["S02", "园区负荷规模", "Frontiers in Energy Research: heavy equipment manufacturing industrial park IES", "论文", "2 km2 industrial park, 50+ facilities, 200 MW capacity, 150 MW peak demand, 1.2 TWh electricity and 0.8 TWh thermal energy annually", "https://www.frontiersin.org/journals/energy-research/articles/10.3389/fenrg.2024.1448362/full", "2026-06-18"],
        ["S03", "气象与风光资源", "NASA POWER Hourly API", "公开数据集", "Ordos point hourly ALLSKY_SFC_SW_DWN, WS50M, T2M, 2024, LST", NASA_URL, "2026-06-18"],
        ["S04", "蒙西分时电价", "锡林郭勒盟发改委: 内蒙古西部电网销售电价和输配电价公示表", "政府", "蒙西峰平谷时段、峰谷比价、尖峰/深谷规则、需量电价", "https://fgw.xlgl.gov.cn/fgw/ywgz/jgsf/jfgs/jggs/2023112710530866099/index.html", "2026-06-18"],
        ["S05", "管道天然气价格", "鄂尔多斯市发改委非居民管道天然气价格通知", "政府", "2026-04-01 至 2026-10-31 非居民终端销售价格 2.952 元/m3", "https://www.ordos.gov.cn/ordosml/szbm/sfzhggwyh_n/202605/t20260528_3900116.html", "2026-06-18"],
        ["S06", "电网排放因子", "生态环境部/国家统计局 2022年电力二氧化碳排放因子", "政府", "内蒙古省级电力平均二氧化碳排放因子 0.6849 kgCO2/kWh", "https://www.mee.gov.cn/xxgk2018/xxgk/xxgk01/202412/t20241226_1099413.html", "2026-06-18"],
        ["S07", "碳价", "生态环境部: 2025年全国碳市场运行情况", "政府", "2025 全年交易均价 62.36 元/tCO2，年底收盘价 74.63 元/tCO2", "https://www.mee.gov.cn/ywgz/ydqhbh/wsqtkz/202601/t20260101_1139528.shtml", "2026-06-18"],
        ["S08", "可再生能源成本", "IRENA Renewable Power Generation Costs in 2024", "国际组织报告", "2024 PV TIC 691 USD/kW, onshore wind TIC 1041 USD/kW, battery storage 192 USD/kWh", "https://www.irena.org/-/media/Files/IRENA/Agency/Publication/2025/Jul/IRENA_TEC_RPGC_in_2024_Summary_2025.pdf", "2026-06-18"],
        ["S09", "储能成本", "CNESA Energy Storage Industry White Paper 2024 Summary", "行业报告", "2023 年底 2h LFP 储能系统平均中标价 0.79 元/Wh，部分集采低于 0.6 元/Wh", "https://en.cnesa.org/s/CNESA-Energy-Storage-Industry-WhitePaper-2024Summary-Version.pdf", "2026-06-18"],
        ["S10", "电解槽参数", "DOE Technical Targets for PEM Electrolysis", "政府/研发目标", "PEM 2022 status: system efficiency 55 kWh/kg H2, uninstalled capital cost 1000 USD/kW, lifetime 40,000 h", "https://www.energy.gov/cmei/fuels/technical-targets-proton-exchange-membrane-electrolysis", "2026-06-18"],
        ["S11", "电解槽效率和成本范围", "IRENA Green Hydrogen Cost Reduction", "国际组织报告", "Alkaline system 50-78 kWh/kg H2 and 500-1000 USD/kW; PEM system 50-83 kWh/kg H2 and 700-1400 USD/kW", "https://www.irena.org/-/media/Files/IRENA/Agency/Publication/2020/Dec/IRENA_Green_hydrogen_cost_2020.pdf", "2026-06-18"],
        ["S12", "中国氢价", "国家能源局: 中国氢能发展报告(2025)", "政府/行业统计", "2024年底生产侧 28.0 元/kg，消费侧 48.6 元/kg；内蒙古等为氢消费重点地区", "https://www.nea.gov.cn/20250430/96022785b3a747248288ad1c57d3a025/83d863317f2f44edb605348e4de40993.pdf", "2026-06-18"],
        ["S13", "热泵成本与COP", "IRENA Heat Pump Costs and Markets", "国际组织报告", "Commercial air-air/air-water heat pump costs and COP values; e.g. 358-627 USD/kWth and COP around 3.9-4.2 in reported cases", "https://www.irena.org/-/media/Files/IRENA/Agency/Publication/2022/Nov/IRENA_Heat_Pumps_Costs_Markets_2022.pdf", "2026-06-18"],
        ["S14", "热泵/锅炉效率", "IEA industrial heat pump chart notes", "国际组织数据", "Assumed efficiencies: coal boilers 84%, natural gas boilers 87%, electric boilers 98%, heat pumps 350%", "https://www.iea.org/data-and-statistics/charts/energy-input-cost-ratio-of-industrial-heat-pumps-relative-to-current-market-benchmarks-in-selected-regions-2024-3", "2026-06-18"],
        ["S15", "燃料电池成本", "DOE Heavy-Duty Fuel Cell System Cost 2023", "政府/研发报告", "275 kW PEM fuel cell system projected about 170 USD/kW at 50k units/year; 25,000 h durability target", "https://www.hydrogen.energy.gov/docs/hydrogenprogramlibraries/pdfs/review24/24004-hd-fuel-cell-system-cost-2023.pdf", "2026-06-18"],
        ["S16", "储氢成本", "Hydrogen Storage and Transport: Technologies and Costs", "论文/报告", "Liquid hydrogen large tank cost 30-50 USD/kgH2; compressed gas storage references included for sensitivity", "https://escholarship.org/content/qt83p5k54m/qt83p5k54m_noSplash_8bb1326c13cfb9aa3d0d376ec26d3e06.pdf", "2026-06-18"],
        ["S17", "天然气排放计算", "国家发改委工业其他行业企业温室气体核算方法与报告指南", "政府指南", "天然气低位发热量、单位热值含碳量、碳氧化率用于推导约 2.16 kgCO2/m3", "https://www.ndrc.gov.cn/xxgk/zcfb/tz/201511/W020190905506438889540.pdf", "2026-06-18"],
        ["S18", "新能源消纳约束", "国家能源局: 电力系统调节能力优化专项行动", "政府", "2025-2027 年全国新能源利用率不低于 90%", "https://www.nea.gov.cn/20250117/7a37e1cc43e4477baa101671b7dc273f/c.html", "2026-06-18"],
        ["S19", "电力市场机制", "RMI 2024 China Power Market Outlook", "研究报告", "I&C price components, capacity pricing, ancillary services and energy-storage market participation context", "https://rmi.org/app/uploads/dlm_uploads/2025/01/final-2024-China-Power-Market-Outlook-10-Key-Trends-for-Market-Players-2501117.pdf", "2026-06-18"],
        ["S20", "备用价值/停电损失", "LBNL outage cost literature review", "研究报告", "Industrial outage value examples include 7.61 USD/kWh in reviewed literature; use as sensitivity anchor", "https://eta-publications.lbl.gov/sites/default/files/erss_manuscript_preprint_0.pdf", "2026-06-18"],
    ]


def parameter_rows() -> list[list[object]]:
    return [
        ["P0", "时序负荷", "园区逐时电负荷", "electric_load_kw", "见 TimeSeries_2024_hourly", "", "", "kW", "逐时", "鄂尔多斯重构", 2024, "S01;S02", "按公开工业园区 150MW峰值/1.2TWh年电量案例重构，不是园区实测", "中"],
        ["P0", "时序负荷", "园区逐时热负荷", "heat_load_kw", "见 TimeSeries_2024_hourly", "", "", "kWth", "逐时", "鄂尔多斯重构", 2024, "S02;S03", "工业过程热基荷 + 温度敏感供热负荷重构", "中"],
        ["P0", "时序负荷", "园区逐时氢负荷", "hydrogen_load_kg", "见 TimeSeries_2024_hourly", "", "", "kg/h", "逐时", "鄂尔多斯重构", 2024, "S01;S12", "按氢能产业消纳场景重构，非实测", "中低"],
        ["P0", "风光资源", "光伏容量因子", "pv_cf", "见 TimeSeries_2024_hourly", 0, 0.95, "0-1", "逐时", "鄂尔多斯", 2024, "S03", "NASA GHI 按 PR=0.82、温度修正推导", "高"],
        ["P0", "风光资源", "风电容量因子", "wind_cf", "见 TimeSeries_2024_hourly", 0, 0.95, "0-1", "逐时", "鄂尔多斯", 2024, "S03", "NASA 50m风速按 cut-in 3/rated 12/cut-out 25 m/s 推导", "中"],
        ["P0", "能源价格", "购电分时电价", "electricity_price_cny_per_kwh", "见 TimeSeries_2024_hourly", 0.1584, 0.8316, "元/kWh", "逐时", "蒙西", 2026, "S04;S19", "以平段 0.45 元/kWh 为建模假设，再按蒙西峰谷比价生成", "中"],
        ["P0", "能源价格", "天然气价格", "gas_price_cny_per_m3", GAS_PRICE, 2.952, 3.142, "元/m3", "固定值", "鄂尔多斯东胜/康巴什", 2026, "S05", "非居民终端销售价；罕台镇价格更高", "高"],
        ["P0", "碳排放", "电网排放因子", "grid_emission_kgco2_per_kwh", GRID_EF_INNER_MONGOLIA, "", "", "kgCO2/kWh", "固定值", "内蒙古", 2022, "S06", "省级电力平均二氧化碳排放因子", "高"],
        ["P0", "碳排放", "天然气排放因子", "gas_emission_kgco2_per_m3", 2.16, 2.14, 2.20, "kgCO2/m3", "固定值", "中国", 2024, "S17", "按低位发热量、含碳量、氧化率推导", "高"],
        ["P0", "设备成本", "光伏投资成本", "pv_capex_cny_per_kw", round(691 * USD_CNY), 3500, 5500, "元/kW", "固定值", "全球/中国适配", 2024, "S08", "IRENA 2024 TIC 691 USD/kW 折算，模型可用中国低成本情景", "中高"],
        ["P0", "设备成本", "风电投资成本", "wind_capex_cny_per_kw", round(1041 * USD_CNY), 4200, 7600, "元/kW", "固定值", "全球/中国适配", 2024, "S08", "IRENA 2024 onshore wind TIC 1041 USD/kW 折算", "中高"],
        ["P0", "设备成本", "电池储能能量成本", "battery_energy_capex_cny_per_kwh", 1000, 600, 1300, "元/kWh", "固定值", "中国", 2024, "S09;S08", "CNESA 系统中标价 0.79元/Wh，IRENA全球储能 192 USD/kWh 交叉校验", "中高"],
        ["P0", "设备成本", "电池储能功率成本", "battery_power_capex_cny_per_kw", 800, 500, 1500, "元/kW", "固定值", "中国", 2024, "S09", "PCS/并网/BOP 估算，建议敏感性分析", "中"],
        ["P0", "设备成本", "碱性电解槽投资成本", "electrolyzer_capex_cny_per_kw", 2200, 1500, 7250, "元/kW", "固定值", "中国/国际", 2024, "S10;S11", "中国低成本设备取低值，PEM/国际系统取高值", "中"],
        ["P0", "设备成本", "储氢投资成本", "hydrogen_storage_capex_cny_per_kg", 3600, 220, 5100, "元/kgH2", "固定值", "国际/中国适配", 2024, "S16", "液氢大罐低，压缩储氢高；园区短周期可做敏感性", "中"],
        ["P0", "设备成本", "燃料电池投资成本", "fuel_cell_capex_cny_per_kw", 6000, 1200, 12000, "元/kW", "固定值", "国际/中国适配", 2024, "S15", "交通用 PEM 成本低于固定式系统，园区备用取中值", "中"],
        ["P0", "设备成本", "热泵投资成本", "heat_pump_capex_cny_per_kwth", 3500, 2600, 4600, "元/kWth", "固定值", "国际/中国适配", 2024, "S13", "IRENA 商业热泵成本折算", "中"],
        ["P0", "设备成本", "燃气锅炉投资成本", "gas_boiler_capex_cny_per_kwth", 500, 300, 800, "元/kWth", "固定值", "工程假设", 2024, "S14", "成本需后续用具体厂家报价替换", "低"],
        ["P0", "设备寿命", "光伏寿命", "pv_lifetime_year", 25, 25, 30, "年", "固定值", "行业通用", 2024, "S08", "容量规划年化用", "中"],
        ["P0", "设备寿命", "风电寿命", "wind_lifetime_year", 20, 20, 25, "年", "固定值", "行业通用", 2024, "S08", "容量规划年化用", "中"],
        ["P0", "设备寿命", "电池寿命", "battery_lifetime_year", 12, 10, 15, "年", "固定值", "行业通用", 2024, "S09", "取决于循环和日历衰减", "中"],
        ["P0", "设备寿命", "电解槽寿命", "electrolyzer_lifetime_year", 15, 8, 20, "年", "固定值", "国际/中国适配", 2024, "S10;S11", "也可用 40,000-80,000 小时约束", "中"],
        ["P0", "设备寿命", "储氢寿命", "hydrogen_storage_lifetime_year", 20, 15, 30, "年", "固定值", "行业通用", 2024, "S16", "压力容器需按检验周期校核", "中"],
        ["P0", "设备寿命", "燃料电池寿命", "fuel_cell_lifetime_year", 10, 5, 15, "年", "固定值", "国际/中国适配", 2024, "S15", "DOE 重卡系统耐久目标 25,000h；固定式可更高", "中"],
        ["P0", "设备效率", "电池充电效率", "battery_charge_efficiency", 0.95, 0.92, 0.97, "-", "固定值", "行业通用", 2024, "S09", "与放电效率组合约 90% 往返", "中"],
        ["P0", "设备效率", "电池放电效率", "battery_discharge_efficiency", 0.95, 0.92, 0.97, "-", "固定值", "行业通用", 2024, "S09", "与充电效率组合约 90% 往返", "中"],
        ["P0", "设备效率", "电解槽耗电", "electrolyzer_kwh_per_kg", 55, 50, 78, "kWh/kgH2", "固定值/可做曲线", "国际/中国适配", 2024, "S10;S11", "PEM 2022 status 55；碱性/PEM系统范围 50-83", "高"],
        ["P0", "设备效率", "燃料电池发电效率", "fuel_cell_kwh_per_kg", 18.5, 16.7, 20.0, "kWh/kgH2", "固定值/可做曲线", "国际/中国适配", 2024, "S15", "按 H2 LHV 33.33 kWh/kg 和 50%-60% 电效率推导", "中"],
        ["P0", "设备效率", "热泵 COP", "heat_pump_cop", 3.5, 2.5, 4.2, "-", "固定值/温度曲线", "国际/中国适配", 2024, "S13;S14", "IEA 工业热泵效率 350%，IRENA 案例 COP 约 3.9-4.2", "中高"],
        ["P0", "设备效率", "燃气锅炉效率", "gas_boiler_efficiency", 0.87, 0.84, 0.92, "-", "固定值", "国际/中国适配", 2024, "S14", "IEA 天然气锅炉效率 87%", "高"],
        ["P0", "设备效率", "天然气低位热值", "gas_lhv_kwh_per_m3", 10.814, 10.5, 11.0, "kWh/m3", "固定值", "中国", 2024, "S17", "38.931 MJ/m3 / 3.6", "高"],
        ["P1", "市场机制", "余电上网价格", "grid_sell_price_cny_per_kwh", 0.2829, 0.20, 0.45, "元/kWh", "固定值", "蒙西", 2025, "S19", "按蒙西煤电基准价/机制电价近似，现货和绿电交易需另行替换", "中"],
        ["P1", "市场机制", "最大上网容量", "max_grid_export_kw", 50000, 0, 100000, "kW", "固定值", "园区假设", 2024, "S02", "按 200MW 接入能力的 25% 设置，可由并网批复替换", "低"],
        ["P1", "市场机制", "最大购电容量", "max_grid_import_kw", 200000, 150000, 250000, "kW", "固定值", "园区假设", 2024, "S02", "公开案例有 200MW capacity 和 150MW peak demand", "中"],
        ["P1", "市场机制", "需量电费", "demand_charge_cny_per_kw_year", 393.6, 374.4, 393.6, "元/kW·年", "固定值", "蒙西", 2026, "S04", "35kV/110kV 31.2元/kW·月，1-10kV 32.8元/kW·月", "高"],
        ["P1", "市场机制", "碳价", "carbon_price_cny_per_tco2", CARBON_PRICE, 50, 100, "元/tCO2", "固定值", "全国碳市场", 2025, "S07", "采用 2025 年全年交易均价", "高"],
        ["P1", "政策约束", "年碳排放上限", "annual_carbon_cap_kgco2", 350000000, 100000000, 800000000, "kgCO2/年", "场景值", "园区假设", 2024, "S01;S06", "零碳园区约束场景，建议按基准排放削减率设置", "低"],
        ["P1", "政策约束", "新能源最低消纳率", "min_renewable_utilization", 0.90, 0.90, 0.98, "%", "固定值/场景", "中国", 2025, "S18", "国家能源局提出 2025-2027 全国新能源利用率不低于90%", "高"],
        ["P1", "氢市场", "外购氢价格", "external_h2_purchase_cny_per_kg", 48.6, 28.0, 56.4, "元/kg", "固定值", "中国", 2024, "S12", "消费侧价格作为兜底外购价，生产侧价格作为低值", "高"],
        ["P1", "氢市场", "售氢价格", "h2_sale_price_cny_per_kg", 45.7, 28.0, 56.4, "元/kg", "固定值", "中国", 2024, "S12", "示范城市群消费侧价作对外售价参考", "中高"],
        ["P1", "运维退化", "电池退化成本", "battery_degradation_cny_per_kwh_throughput", 0.10, 0.06, 0.20, "元/kWh throughput", "固定值", "推导", 2024, "S09", "按储能成本/可用循环吞吐估算", "中"],
        ["P1", "运维退化", "光伏固定运维成本", "pv_fixed_om_cny_per_kw_year", 60, 40, 90, "元/kW·年", "固定值", "行业通用", 2024, "S08", "约为 CAPEX 1%-2%", "中"],
        ["P1", "运维退化", "风电固定运维成本", "wind_fixed_om_cny_per_kw_year", 140, 90, 220, "元/kW·年", "固定值", "行业通用", 2024, "S08", "约为 CAPEX 2%-3%", "中"],
        ["P1", "运维退化", "电解槽可变运维", "electrolyzer_var_om_cny_per_kg", 1.0, 0.5, 2.5, "元/kgH2", "固定值", "工程假设", 2024, "S10;S11", "不含电费", "低"],
        ["P1", "运维退化", "燃料电池可变运维", "fuel_cell_var_om_cny_per_kwh", 0.08, 0.03, 0.20, "元/kWh", "固定值", "工程假设", 2024, "S15", "备电/低利用小时场景需敏感性", "低"],
        ["P2", "细化曲线", "热泵COP冬季", "heat_pump_cop_winter", 2.6, 2.2, 3.2, "-", "季节", "鄂尔多斯假设", 2024, "S03;S13;S14", "低温空气源热泵按气温下修", "中"],
        ["P2", "细化曲线", "热泵COP夏季", "heat_pump_cop_summer", 4.0, 3.5, 4.5, "-", "季节", "鄂尔多斯假设", 2024, "S13;S14", "夏季/过渡季低温差", "中"],
        ["P2", "细化曲线", "电解槽最小负荷率", "electrolyzer_min_load", 0.15, 0.05, 0.30, "%", "固定值", "国际/中国适配", 2024, "S11", "碱性 15%-100%，PEM 可更宽", "中"],
        ["P2", "细化曲线", "燃料电池最小负荷率", "fuel_cell_min_load", 0.20, 0.10, 0.30, "%", "固定值", "工程假设", 2024, "S15", "用于避免小负荷低效运行", "低"],
        ["P2", "储氢", "储氢日损耗率", "hydrogen_storage_loss_per_day", 0.001, 0.0, 0.005, "%/day", "固定值", "工程假设", 2024, "S16", "压缩氢短周期可近似为0，液氢需考虑蒸发", "中低"],
        ["P2", "备用价值", "燃料电池备用容量价值", "fuel_cell_backup_value_cny_per_kw_year", 300, 100, 1000, "元/kW·年", "场景值", "推导", 2024, "S20", "由停电损失和备用替代成本推导，需实测可靠性校准", "低"],
        ["P2", "备用价值", "停电损失", "value_of_lost_load_cny_per_kwh", 55.2, 15, 350, "元/kWh", "场景值", "文献折算", 2024, "S20", "7.61 USD/kWh * 7.25，工业差异极大", "中低"],
        ["P2", "备用价值", "年停电小时数", "annual_outage_hours", 2.0, 0.5, 10.0, "h/年", "场景值", "园区假设", 2024, "S20", "需用当地供电可靠性报告替换", "低"],
        ["P2", "财务", "折现率", "discount_rate", 0.06, 0.04, 0.08, "%", "固定值", "财务假设", 2024, "S08;S19", "容量规划年化成本用", "中"],
    ]


def write_table(ws, headers: list[str], rows: list[list[object]], name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    end_row = len(rows) + 1
    end_col = len(headers)
    ref = f"A1:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:200]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def add_readme(wb: Workbook, ts_count: int) -> None:
    ws = wb.active
    ws.title = "README_说明"
    rows = [
        ["文件名", "codex数据汇总表.xlsx"],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["项目背景", "鄂尔多斯零碳产业园公开案例 + 公开数据集 + 公开论文/报告参数 + 工程假设重构"],
        ["重要声明", "本文件未引用项目现有 Excel 数据；负荷曲线为公开案例尺度校准后的重构数据，非园区实测。"],
        ["时序覆盖", f"2024 全年逐时 {ts_count} 条；另含 3 个典型日共 72 条。"],
        ["坐标", f"鄂尔多斯代表点 latitude={LATITUDE}, longitude={LONGITUDE}"],
        ["气象来源", "NASA POWER Hourly API: ALLSKY_SFC_SW_DWN, WS50M, T2M"],
        ["电价说明", "蒙西分时政策给出峰谷时段和比价；平段电价采用 0.45 元/kWh 建模假设，可替换为企业交易合同价格。"],
        ["推荐使用", "模型输入优先读取 TimeSeries_2024_hourly、Parameters_推荐参数、TOU_分时电价、Sources_来源索引。"],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100


def add_tou_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("TOU_分时电价")
    headers = ["季节", "月份", "时段", "小时范围", "峰谷系数", "模型电价_元每kWh", "政策依据", "备注"]
    rows = [
        ["大风季", "1-5,9-12", "峰", "06:00-08:00;18:00-22:00", 1.68, round(BASE_ELECTRICITY_PRICE * 1.68, 4), "S04", "峰平谷 1.68:1:0.48"],
        ["大风季", "1-5,9-12", "平", "04:00-06:00;08:00-11:00;16:00-18:00;22:00-24:00", 1.00, BASE_ELECTRICITY_PRICE, "S04", "平段为建模基准价"],
        ["大风季", "1-5,9-12", "谷", "00:00-04:00;11:00-16:00", 0.48, round(BASE_ELECTRICITY_PRICE * 0.48, 4), "S04", "谷段下浮52%"],
        ["小风季", "6-8", "峰", "06:00-08:00;18:00-22:00", 1.54, round(BASE_ELECTRICITY_PRICE * 1.54, 4), "S04", "峰平谷 1.54:1:0.44"],
        ["小风季", "6-8", "平", "00:00-06:00;08:00-11:00;16:00-18:00;22:00-24:00", 1.00, BASE_ELECTRICITY_PRICE, "S04", "平段为建模基准价"],
        ["小风季", "6-8", "谷", "11:00-16:00", 0.44, round(BASE_ELECTRICITY_PRICE * 0.44, 4), "S04", "谷段下浮56%"],
        ["小风季", "6-8", "尖峰", "19:00-21:00", 1.848, round(BASE_ELECTRICITY_PRICE * 1.54 * 1.20, 4), "S04", "尖峰在峰段基础上上浮20%"],
        ["小风季", "6-8", "深谷", "13:00-15:00", 0.352, round(BASE_ELECTRICITY_PRICE * 0.44 * 0.80, 4), "S04", "深谷在谷段基础上下浮20%"],
        ["全年", "1-12", "需量电费", "最大需量", "", 393.6, "S04", "32.8 元/kW·月折年化"],
    ]
    write_table(ws, headers, rows, "TOUTable")
    style_sheet(ws)


def build_time_series(nasa: dict) -> list[list[object]]:
    params = nasa["properties"]["parameter"]
    keys = sorted(params["ALLSKY_SFC_SW_DWN"].keys())
    rows = []
    for key in keys:
        dt = datetime.strptime(key, "%Y%m%d%H")
        ghi = float(params["ALLSKY_SFC_SW_DWN"][key])
        ws50m = float(params["WS50M"][key])
        temp = float(params["T2M"][key])
        pv_cf = pv_capacity_factor(ghi, temp)
        wind_cf = wind_capacity_factor(ws50m)
        period, price = tou_period_and_price(dt)
        electric, heat, hydrogen = reconstructed_loads(dt, temp)
        rows.append([
            dt.strftime("%Y-%m-%d %H:00"),
            dt.date().isoformat(),
            dt.hour,
            season_name(dt.month),
            electric,
            heat,
            hydrogen,
            pv_cf,
            wind_cf,
            ghi,
            ws50m,
            temp,
            period,
            price,
            GAS_PRICE,
            GRID_EF_INNER_MONGOLIA,
            CARBON_PRICE,
            "S03",
            "S01;S02;S12",
            "NASA原始气象 + 公开案例尺度校准负荷重构",
        ])
    return rows


def add_summary(wb: Workbook, ts_rows: list[list[object]]) -> None:
    ws = wb.create_sheet("Coverage_覆盖检查", 0)
    headers = ["检查项", "结果", "说明"]
    electric = [r[4] for r in ts_rows]
    heat = [r[5] for r in ts_rows]
    h2 = [r[6] for r in ts_rows]
    pv = [r[7] for r in ts_rows]
    wind = [r[8] for r in ts_rows]
    rows = [
        ["全年小时数", len(ts_rows), "2024 为闰年，完整应为 8784 小时"],
        ["电负荷峰值_kW", round(max(electric), 1), "公开工业园区案例峰值 150MW 量级校准"],
        ["电负荷年电量_MWh", round(sum(electric) / 1000, 1), "模型重构值"],
        ["热负荷峰值_kWth", round(max(heat), 1), "按温度敏感过程热重构"],
        ["热负荷年热量_MWhth", round(sum(heat) / 1000, 1), "模型重构值"],
        ["氢负荷峰值_kg/h", round(max(h2), 2), "按氢能产业消纳场景重构"],
        ["氢负荷年需求_t", round(sum(h2) / 1000, 1), "模型重构值"],
        ["光伏容量因子均值", round(sum(pv) / len(pv), 4), "NASA GHI 推导"],
        ["风电容量因子均值", round(sum(wind) / len(wind), 4), "NASA 50m 风速推导"],
        ["P0核心项覆盖", "完成", "负荷、风光、电价、设备成本/寿命/效率、排放因子、气价均已覆盖"],
        ["P1增强项覆盖", "完成", "上网电价、购售电容量、需量电费、碳价、氢价、消纳率、退化/运维均已覆盖"],
        ["P2细化项覆盖", "完成/假设", "COP曲线、负荷率、储氢损耗、备用价值、停电损失等提供建模假设和来源"],
    ]
    write_table(ws, headers, rows, "CoverageTable")
    style_sheet(ws)


def main() -> None:
    nasa = get_nasa_power()
    ts_rows = build_time_series(nasa)

    wb = Workbook()
    add_readme(wb, len(ts_rows))
    add_summary(wb, ts_rows)

    sources_ws = wb.create_sheet("Sources_来源索引")
    write_table(
        sources_ws,
        ["source_id", "类别", "来源名称", "来源类型", "可用数据/用途", "url", "检索日期"],
        source_rows(),
        "SourcesTable",
    )
    style_sheet(sources_ws)

    params_ws = wb.create_sheet("Parameters_推荐参数")
    write_table(
        params_ws,
        ["优先级", "类别", "数据名称", "字段名", "推荐值", "低值", "高值", "单位", "时间尺度", "地区", "年份", "来源ID", "推导/备注", "可信度"],
        parameter_rows(),
        "ParametersTable",
    )
    style_sheet(params_ws)

    tou_ws_exists = "TOU_分时电价" in wb.sheetnames
    if not tou_ws_exists:
        add_tou_sheet(wb)

    ts_ws = wb.create_sheet("TimeSeries_2024_hourly")
    ts_headers = [
        "datetime",
        "date",
        "hour",
        "season",
        "electric_load_kw",
        "heat_load_kw",
        "hydrogen_load_kg",
        "pv_cf",
        "wind_cf",
        "ghi_wh_m2",
        "wind_speed_50m_m_s",
        "temperature_c",
        "tou_period",
        "electricity_price_cny_per_kwh",
        "gas_price_cny_per_m3",
        "grid_emission_kgco2_per_kwh",
        "carbon_price_cny_per_tco2",
        "weather_source_id",
        "load_source_ids",
        "method_note",
    ]
    write_table(ts_ws, ts_headers, ts_rows, "TimeSeriesTable")
    style_sheet(ts_ws)
    for col in ["E", "F", "G", "H", "I", "J", "K", "L", "N", "O", "P", "Q"]:
        for cell in ts_ws[col][1:]:
            cell.number_format = "0.0000" if col in ("H", "I", "N", "O", "P", "Q") else "0.0"

    typical_dates = {"2024-01-15": "典型冬季日", "2024-04-15": "典型过渡季日", "2024-07-15": "典型夏季日"}
    typ_rows = [r + [typical_dates[r[1]]] for r in ts_rows if r[1] in typical_dates]
    typ_ws = wb.create_sheet("TypicalDays_72h")
    write_table(typ_ws, ts_headers + ["typical_day_type"], typ_rows, "TypicalDaysTable")
    style_sheet(typ_ws)

    assumptions_ws = wb.create_sheet("Assumptions_推导公式")
    assumption_rows = [
        ["光伏容量因子", "pv_cf = min(0.95, GHI_Wh_m2/1000 * 0.82 * temp_factor)", "PR=0.82；组件温度 Tmodule=T2M+20*(GHI/800)；高于25C按 -0.4%/C 修正", "S03"],
        ["风电容量因子", "cut-in 3m/s, rated 12m/s, cut-out 25m/s; between cut-in and rated uses cubic curve", "用50m风速直接推导，未做轮毂高度和空气密度修正", "S03"],
        ["电负荷重构", "150MW峰值约束，高基荷三班制曲线", "公开制造业园区案例 150MW峰值/1.2TWh年用电作为尺度锚点", "S02"],
        ["热负荷重构", "process_heat + max(0,18-T2M)*1900 + summer_hot_water", "用于模拟工业过程热 + 北方采暖季热负荷", "S02;S03"],
        ["氢负荷重构", "约 780-1100 kg/h，日间物流/生产上浮", "根据氢能产业消纳场景给出连续用氢曲线，非实测", "S01;S12"],
        ["购电价格", "平段假设0.45元/kWh，按蒙西峰谷比价生成逐时电价", "用户拿到企业交易合同后应替换平段价格", "S04;S19"],
        ["天然气排放因子", "38.931MJ/m3 * 15.3e-3 tC/GJ * 99% * 44/12 = 2.16kgCO2/m3", "按中国核算指南默认参数推导", "S17"],
        ["燃料电池发电效率", "33.33 kWh/kgH2 * 55% = 18.33 kWh/kgH2", "推荐值取18.5", "S15"],
    ]
    write_table(assumptions_ws, ["项目", "公式/方法", "说明", "来源ID"], assumption_rows, "AssumptionsTable")
    style_sheet(assumptions_ws)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = ws.freeze_panes or "A2"

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
